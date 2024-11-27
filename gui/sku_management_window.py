import os
import random
import string
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QLabel, QComboBox, QPushButton, QFileDialog,
    QMessageBox, QTableWidget, QTableWidgetItem, QHBoxLayout, QLineEdit
)
from PyQt5.QtCore import Qt
from database.database_handler import DatabaseHandler
import openpyxl
from utils.validators import validate_vendor_name


class SkuManagementWindow(QDialog):
    """Window for managing SKUs for a given vendor."""
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Manage SKUs")
        self.resize(900, 600)

        # Initialize database handler
        self.db_handler = DatabaseHandler()

        # Main layout
        layout = QVBoxLayout()

        # Vendor dropdown
        self.vendor_selector = QComboBox()
        self.vendor_selector.addItem("Select Vendor")
        self.load_existing_vendors()
        layout.addWidget(QLabel("Select Vendor:"))
        layout.addWidget(self.vendor_selector)

        # Input field for SKU count
        self.sku_count_input = QLineEdit()
        self.sku_count_input.setPlaceholderText("Enter the number of SKUs to generate")
        layout.addWidget(QLabel("Number of SKUs to generate:"))
        layout.addWidget(self.sku_count_input)

        # Buttons for SKU operations
        button_layout = QHBoxLayout()
        self.export_button = QPushButton("Export SKUs")
        self.export_button.clicked.connect(self.export_skus)
        self.import_button = QPushButton("Import SKUs")
        self.import_button.clicked.connect(self.import_skus)
        self.generate_button = QPushButton("Generate SKUs")
        self.generate_button.clicked.connect(self.generate_skus)
        button_layout.addWidget(self.export_button)
        button_layout.addWidget(self.import_button)
        button_layout.addWidget(self.generate_button)
        layout.addLayout(button_layout)

        # SKU Table
        self.sku_table = QTableWidget()
        self.sku_table.setColumnCount(4)
        self.sku_table.setHorizontalHeaderLabels(["SKU", "Status", "Old SKU", "Report"])
        layout.addWidget(self.sku_table)

        self.setLayout(layout)

    def load_existing_vendors(self):
        """Load vendor names into the dropdown."""
        self.vendor_selector.clear()
        vendors = self.db_handler.get_collection("vendors").find()
        for vendor in vendors:
            self.vendor_selector.addItem(vendor["name"])

    def export_skus(self):
        """Export SKUs for the selected vendor to a spreadsheet."""
        vendor_name = self.vendor_selector.currentText()
        if not vendor_name:
            QMessageBox.warning(self, "Error", "Please select a vendor.")
            return

        # Fetch products for the vendor
        products = self.db_handler.get_collection("all_products").find({"vendor.vendor_name": vendor_name})

        # Create a spreadsheet
        file_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
        if not file_path:
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["vendor_name", "sku", "sku_type", "sku_status", "old_sku", "report"])  # Headers

        for product in products:
            sku_info = product.get("sku_info", {})
            sheet.append([
                vendor_name,
                sku_info.get("sku", ""),
                sku_info.get("sku_type", ""),
                sku_info.get("sku_status", ""),
                sku_info.get("old_sku", ""),
                ""  # Empty report column
            ])

        workbook.save(file_path)
        QMessageBox.information(self, "Success", f"SKUs exported to {file_path}")

    def import_skus(self):
        """Import SKUs from a spreadsheet, update the database, and generate a report file."""
        file_path, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Excel Files (*.xlsx)")
        if not file_path:
            return

        # Load the spreadsheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Define expected headers
        EXPECTED_HEADERS = ["vendor_name", "sku", "sku_type", "sku_status", "old_sku", "report"]

        # Validate headers
        headers = [cell.value for cell in sheet[1]]
        if headers[:len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
            QMessageBox.critical(self, "Error", "Spreadsheet headers do not match the expected format.")
            return

        print(f"Importing SKUs from file: {file_path}")

        # Process only designated columns
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS), values_only=True),
                                      start=2):
            print(f"Row {row_idx}: {row}")  # Debugging
            vendor_name, sku, sku_type, sku_status, old_sku, report = row

            # Handle None values
            vendor_name = vendor_name or ""
            sku = sku or ""
            sku_type = sku_type or ""
            sku_status = sku_status or ""
            old_sku = old_sku if old_sku else "NA"
            report = report or ""

            try:
                validate_vendor_name(vendor_name)
                print(f"Row {row_idx}: Vendor '{vendor_name}' validated successfully.")

                product_data = {
                    "sku_info": {
                        "sku": sku,
                        "sku_type": sku_type,
                        "sku_status": sku_status,
                        "old_sku": old_sku
                    },
                    "vendor": {"vendor_name": vendor_name}
                }

                collection = self.db_handler.get_collection("all_products")

                # Perform actions based on sku_status
                if sku_status == "replace":
                    # Replace SKU
                    old_product = collection.find_one({"vendor.vendor_name": vendor_name, "sku_info.sku": old_sku})
                    if old_product:
                        collection.update_one(
                            {"_id": old_product["_id"]},
                            {"$set": {"sku_info.old_sku": old_sku, "sku_info.sku": sku}}
                        )
                        report = f"Replaced the SKU '{old_sku}' with '{sku}'"
                    else:
                        report = f"Old SKU '{old_sku}' not found. Replacement failed."

                elif sku_status == "discontinued":
                    # Mark as discontinued
                    existing_product = collection.find_one({"vendor.vendor_name": vendor_name, "sku_info.sku": sku})
                    if existing_product:
                        collection.update_one(
                            {"_id": existing_product["_id"]},
                            {"$set": {"sku_info.sku_status": "discontinued"}}
                        )
                        report = f"Marked SKU '{sku}' as discontinued."
                    else:
                        report = f"SKU '{sku}' not found. Cannot mark as discontinued."

                else:  # Handle "current" or other cases
                    result = collection.update_one(
                        {"sku_info.sku": sku, "vendor.vendor_name": vendor_name},
                        {"$set": product_data},
                        upsert=True
                    )

                    if result.upserted_id:
                        report = f"Created a new document for SKU '{sku}'."
                    elif result.modified_count > 0:
                        report = f"Updated existing document for SKU '{sku}'."
                    else:
                        report = f"SKU '{sku}' already exists. No changes made."

            except ValueError as e:
                print(f"Row {row_idx}: Validation error: {e}")
                report = f"Validation error: {e}"
            except Exception as e:
                print(f"Row {row_idx}: Unexpected error: {e}")
                report = f"Unexpected error: {e}"

            # Write the report back to the spreadsheet
            sheet.cell(row=row_idx, column=6).value = report

        # Save the updated spreadsheet as a report file
        report_file_path = file_path.replace(".xlsx", "_import_report.xlsx")
        workbook.save(report_file_path)

        QMessageBox.information(self, "Success",
                                f"SKUs imported and updated successfully.\nReport saved to: {report_file_path}")

    def generate_skus(self):
        """Generate SKUs for the selected vendor."""
        try:
            # Validate vendor selection
            vendor_name = self.vendor_selector.currentText()
            if vendor_name == "Select Vendor":
                raise ValueError("Please select a valid vendor.")
            validate_vendor_name(vendor_name)

            # Validate SKU count
            sku_count_text = self.sku_count_input.text()
            if not sku_count_text.isdigit() or int(sku_count_text) <= 0:
                raise ValueError("Please enter a positive number for SKU count.")
            sku_count = int(sku_count_text)

            # Generate unique SKUs
            generated_skus = []
            collection = self.db_handler.get_collection("all_products")
            for _ in range(sku_count):
                while True:
                    # Generate a unique SKU
                    new_sku = "GEN-" + "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
                    if not collection.find_one({"sku_info.sku": new_sku}):
                        break

                # Insert the new SKU into the database
                product_data = {
                    "sku_info": {
                        "sku": new_sku,
                        "sku_type": "generated",
                        "sku_status": "current",
                        "old_sku": "NA"
                    },
                    "vendor": {"vendor_name": vendor_name}
                }
                collection.insert_one(product_data)

                # Add to the list of generated SKUs
                generated_skus.append({
                    "vendor_name": vendor_name,
                    "sku": new_sku,
                    "sku_type": "generated",
                    "sku_status": "current",
                    "old_sku": "NA",
                    "report": "Generated new SKU"
                })

            # Save to a report
            report_file_path = self.save_sku_report(vendor_name, generated_skus)
            QMessageBox.information(self, "Success",
                                    f"Generated {sku_count} SKUs.\nReport saved to: {report_file_path}")
        except ValueError as e:
            QMessageBox.warning(self, "Error", str(e))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An unexpected error occurred: {e}")

    def save_sku_report(self, vendor_name, generated_skus):
        """Save the generated SKUs to a report file."""
        # Create the report file path
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Report", f"{vendor_name}_generated_skus.xlsx",
                                                   "Excel Files (*.xlsx)")
        if not file_path:
            raise ValueError("No file path selected for saving the report.")

        # Create workbook and add data
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Generated SKUs"

        # Write headers
        headers = ["vendor_name", "sku", "sku_type", "sku_status", "old_sku", "report"]
        sheet.append(headers)

        # Write SKU data
        for sku_data in generated_skus:
            sheet.append([
                sku_data["vendor_name"],
                sku_data["sku"],
                sku_data["sku_type"],
                sku_data["sku_status"],
                sku_data["old_sku"],
                sku_data["report"]
            ])

        # Save workbook
        workbook.save(file_path)
        return file_path


